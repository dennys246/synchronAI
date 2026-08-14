"""Rebuild the CARE second-by-second synchrony labels with correct time decoding.

The shipped labels.csv was built by src/synchronai/data/preprocessing/raw_to_csv.py,
whose _parse_time_to_seconds stringifies openpyxl datetime.time cells ("14:44" on
screen -> time(14, 44) -> "14:44:00") and reads them as HH:MM:SS -> 53040 s. The
coding sheets actually store one-minute date-serial increments (60/86400 per row,
numFmtId 20 "h:mm"), so the display "14:44" is the coder's MM:SS and the true value
is 14*60+44 = 884 s -- every affected row is exactly 60x inflated. labels_filtered.csv
DROPPED those rows (~27k seconds); this script repairs them instead.

Time-cell decoding (see also parse_mmss in scripts/build_new_study_labels.py):
- datetime.time(H, M, S)  -> round((H*3600 + M*60 + S)/60)   [serial*1440: H:MM == MM:SS]
- datetime.datetime       -> the >=24:00 date-serial rollover; decoded via the
                             Excel epoch and the same *1440 rule, and flagged.
- "MM:SS" strings         -> M*60+S (typed by coder; leading spaces stripped)
- fractional float in (0,2) -> bare date-serial, round(v*1440), flagged
- integer                 -> already seconds
Sentinel rows (start/END/total) fail time parsing and drop out.

Per file, the serial interpretation is cross-checked against the strong prior that
per-second coding advances by ~1 s per row; files whose decoded times do not step
by 1 s are flagged in the report rather than silently trusted.

Session combination mirrors the original raw_to_csv semantics (alphabetical file
order, last-wins on per-second conflicts) so the only intended change is the time
decode. Output schema is identical to labels.csv:
    video_path,second,label,subject_id,session

Usage:
    python scripts/build_care_repaired_labels.py --out data/labels_care_repaired.csv \
        [--compare data/labels_filtered.csv] [--limit 5] [--no-ffprobe]
"""

from __future__ import annotations

import argparse
import collections
import datetime
import glob
import os
import re
import shutil
import subprocess

import openpyxl

LOCAL_ROOT = "/Volumes/perlmansusan/Active/moochie"
CLUSTER_ROOT = "/storage1/fs1/perlmansusan/Active/moochie"
DATA_ROOT = next(
    (r for r in (LOCAL_ROOT, CLUSTER_ROOT) if os.path.isdir(f"{r}/study_data")),
    LOCAL_ROOT,
)
STUDY_DATA = f"{DATA_ROOT}/study_data"
SBYS_ROOT = (f"{STUDY_DATA}/CARE/synchrony_coding/second by second/"
             "SbyS synchrony coding participants")
VIDEO_ROOT = f"{STUDY_DATA}/CARE/video_data"
LABEL_ENCODING = {"a": 0, "s": 1}
_EXCEL_EPOCH = datetime.datetime(1899, 12, 30)


def decode_time_cell(val, minute_serial: bool) -> tuple[int | None, str]:
    """Coding Time cell -> (seconds, branch tag). None when not a time.

    Date-serial cells (time/datetime/fractional float) are ambiguous: most
    sheets increment 60/86400 per row (display "h:mm" read as MM:SS ->
    seconds = serial*1440, minute_serial=True), but a minority (e.g. eleven
    MagnetTiles files) increment 1/86400 per row, i.e. the stored time-of-day
    IS the elapsed MM:SS (seconds = serial*86400, minute_serial=False).
    The caller picks the interpretation per file via the 1-second-step prior.
    """
    if val is None or isinstance(val, bool):
        return None, "empty"
    if isinstance(val, datetime.datetime):
        # >=24:00 rollover: the typed value became a date-serial past one day.
        serial = (val - _EXCEL_EPOCH).total_seconds() / 86400.0
        return round(serial * (1440 if minute_serial else 86400)), "serial-datetime"
    if isinstance(val, datetime.time):
        # Summing to seconds first keeps 14:43:59.999997 (accumulated float
        # error from chained +60/86400 formulas) from rounding down to 883.
        total = val.hour * 3600 + val.minute * 60 + val.second + val.microsecond / 1e6
        return round(total / 60) if minute_serial else round(total), "serial-time"
    if isinstance(val, (int, float)):
        if 0 < val < 2 and abs(val - round(val)) > 1e-9:
            return round(val * (1440 if minute_serial else 86400)), "serial-float"
        return int(val), "numeric"
    s = str(val).strip()
    if ":" in s:
        p = s.split(":")
        try:
            if len(p) == 2:
                return int(p[0]) * 60 + int(float(p[1])), "string-mmss"
            if len(p) == 3:
                # Not expected in these sheets; decoded as H:MM:SS and flagged.
                return int(p[0]) * 3600 + int(p[1]) * 60 + int(float(p[2])), "string-hmmss"
        except (ValueError, TypeError):
            return None, "unparsed"
        return None, "unparsed"
    try:
        return int(float(s)), "numeric"
    except (ValueError, TypeError):
        return None, "unparsed"


def read_coding(path: str, branch_counts: collections.Counter,
                second_serial_files: list[str]) -> list[tuple[int, int]]:
    """(second, label) rows from one Time/Code/Notes sheet, in sheet order.

    Decodes the sheet under both serial interpretations and keeps the one
    whose consecutive times step by 1 s more often (per-second coding is
    contiguous, so the right decode scores near 1.0 and the wrong one ~0.02).
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        print(f"    ! failed to open {os.path.basename(path)}: {e}")
        return []
    ws = wb[wb.sheetnames[0]]
    raw: list[tuple[object, int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or row[1] is None:
            continue
        code = str(row[1]).strip().lower()
        if code in LABEL_ENCODING:
            raw.append((row[0], LABEL_ENCODING[code]))
    wb.close()

    def decode_all(minute_serial: bool):
        out, branches = [], collections.Counter()
        for val, lab in raw:
            sec, branch = decode_time_cell(val, minute_serial)
            branches[branch] += 1
            if sec is not None and sec >= 0:
                out.append((sec, lab))
        return out, branches

    minute_rows, minute_branches = decode_all(True)
    if any(b.startswith("serial") for b in minute_branches):
        second_rows, second_branches = decode_all(False)
        if step_fraction([s for s, _ in second_rows]) > step_fraction(
                [s for s, _ in minute_rows]):
            second_serial_files.append(os.path.basename(path))
            branch_counts.update(second_branches)
            branch_counts["files-second-serial"] += 1
            return second_rows
    branch_counts.update(minute_branches)
    return minute_rows


def step_fraction(secs: list[int]) -> float:
    """Fraction of consecutive decoded times that advance by exactly 1 s --
    the per-second-coding continuity prior used to catch a wrong decode."""
    if len(secs) < 2:
        return 1.0
    steps = [b - a for a, b in zip(secs, secs[1:])]
    return sum(1 for d in steps if d == 1) / len(steps)


def discover_sessions() -> list[tuple[str, str, list[str]]]:
    """(subject_id, session, xlsx files) for every populated coding dir."""
    sessions = []
    for sub in sorted(os.listdir(SBYS_ROOT)):
        if not re.fullmatch(r"\d{5}", sub):
            continue
        sub_dir = os.path.join(SBYS_ROOT, sub)
        if not os.path.isdir(sub_dir):
            continue
        for sess in sorted(os.listdir(sub_dir)):
            sess_dir = os.path.join(sub_dir, sess)
            if sess.startswith(".") or not os.path.isdir(sess_dir):
                continue
            files = [
                f for f in sorted(glob.glob(os.path.join(sess_dir, "*.xlsx")))
                if not os.path.basename(f).startswith(("~$", "._"))
            ]
            if files:
                sessions.append((sub, sess, files))
    return sessions


def resolve_video(subject_id: str, session: str) -> str | None:
    """Mirror raw_to_csv.resolve_video_path: exact name, then glob fallback."""
    d = f"{VIDEO_ROOT}/{subject_id[:4]}/{session}"
    if not os.path.isdir(d):
        return None
    exact = f"{d}/{subject_id}_{session}_DB-DOS.mp4"
    if os.path.exists(exact):
        return exact
    for pat in (f"{subject_id}_{session}_DB-DOS.*", f"{subject_id}_{session}_*.mp4",
                f"{subject_id}_{session}_*.MP4", f"{subject_id}_{session}_*.avi"):
        m = [f for f in glob.glob(os.path.join(d, pat))
             if not os.path.basename(f).startswith("._")]
        if m:
            return sorted(m)[0]
    return None


def ffprobe_duration(path: str) -> float | None:
    try:
        out = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=nw=1:nk=1", path],
            capture_output=True, text=True, timeout=120,
        )
        return float(out.stdout.strip())
    except Exception:  # noqa: BLE001
        return None


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="data/labels_care_repaired.csv")
    ap.add_argument("--compare", default="data/labels_filtered.csv",
                    help="prior filtered labels file to report recovery against")
    ap.add_argument("--limit", type=int, default=0,
                    help="dry-run: only process the first N sessions")
    ap.add_argument("--no-ffprobe", action="store_true")
    args = ap.parse_args()

    sessions = discover_sessions()
    if not sessions:
        raise SystemExit(f"ERROR: found 0 coding sessions under {SBYS_ROOT}. "
                         f"Wrong data root? Refusing to write {args.out}.")
    if args.limit:
        sessions = sessions[: args.limit]
    if not args.no_ffprobe and shutil.which("ffprobe") is None:
        print("WARNING: ffprobe not found -- out-of-range filtering DISABLED. "
              "Pass --no-ffprobe to acknowledge, or re-run where it exists.")

    print(f"=== CARE per-second repair: {len(sessions)} sessions under\n    {SBYS_ROOT}")
    branch_counts: collections.Counter = collections.Counter()
    rows_out: list[tuple[str, int, int, str, str]] = []
    per_session: dict[tuple[str, str], int] = {}
    n_novideo = n_conflict_secs = 0
    low_step_files: list[tuple[str, float]] = []
    second_serial_files: list[str] = []

    for sub, sess, files in sessions:
        merged: dict[int, int] = {}
        conflicts = 0
        for f in files:  # alphabetical; last file wins on conflicts (as before)
            rows = read_coding(f, branch_counts, second_serial_files)
            secs = [s for s, _ in rows]
            frac = step_fraction(secs)
            if len(secs) >= 10 and frac < 0.8:
                low_step_files.append((os.path.relpath(f, SBYS_ROOT), frac))
            for sec, lab in rows:
                if sec in merged and merged[sec] != lab:
                    conflicts += 1
                merged[sec] = lab
        n_conflict_secs += conflicts
        vid = resolve_video(sub, sess)
        if vid is None:
            print(f"  {sub}/{sess}: {len(merged):5d} sec -> NO VIDEO (skipped)")
            n_novideo += 1
            continue
        dur = None if args.no_ffprobe else ffprobe_duration(vid)
        oob = 0
        vid_cluster = vid.replace(DATA_ROOT, CLUSTER_ROOT, 1)
        kept = 0
        for sec in sorted(merged):
            if dur is not None and sec + 1 > dur:
                oob += 1
                continue
            rows_out.append((vid_cluster, sec, merged[sec], sub, sess))
            kept += 1
        per_session[(sub, sess)] = kept
        durs = f"{dur:.0f}s" if dur else "n/a"
        extra = (f" OOB={oob}" if oob else "") + (f" conflicts={conflicts}" if conflicts else "")
        print(f"  {sub}/{sess}: {kept:5d} sec kept ({len(files)} files, vid_dur={durs}{extra})")

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    with open(args.out, "w") as fh:
        fh.write("video_path,second,label,subject_id,session\n")
        for vp, sec, lab, s, sess in rows_out:
            fh.write(f"{vp},{sec},{lab},{s},{sess}\n")

    total = len(rows_out)
    dist = collections.Counter(lab for _, _, lab, _, _ in rows_out)
    print(f"\n  sessions kept: {len(per_session)}   no-video: {n_novideo}")
    print(f"  time-cell branches: {dict(branch_counts)}")
    if second_serial_files:
        print(f"  {len(second_serial_files)} file(s) decoded as one-second-per-row "
              f"serials (literal MM:SS): {', '.join(second_serial_files)}")
    if low_step_files:
        print(f"  WARNING: {len(low_step_files)} file(s) with <80% 1-second steps "
              f"(possible wrong decode -- inspect before trusting):")
        for name, frac in low_step_files[:20]:
            print(f"    {frac:.2f}  {name}")
    if total:
        print(f"  labeled seconds: {total}  (async={dist[0]}  sync={dist[1]})")
    print(f"  wrote {args.out}")

    if args.compare and os.path.exists(args.compare) and not args.limit:
        old: dict[tuple[str, str], int] = collections.Counter()
        with open(args.compare) as fh:
            next(fh)
            for line in fh:
                p = line.rstrip("\n").split(",")
                old[(p[3], p[4])] += 1
        old_total = sum(old.values())
        common = set(per_session) & set(old)
        rec_common = sum(per_session[k] - old[k] for k in common)
        new_sessions = set(per_session) - set(old)
        new_secs = sum(per_session[k] for k in new_sessions)
        print(f"\n  --- recovery vs {args.compare} ---")
        print(f"  filtered file: {old_total} seconds over {len(old)} sessions")
        print(f"  repaired file: {total} seconds over {len(per_session)} sessions")
        print(f"  recovered within the {len(common)} common sessions: {rec_common:+d} seconds")
        print(f"  sessions absent from filtered file: {len(new_sessions)} "
              f"(+{new_secs} seconds)")


if __name__ == "__main__":
    main()
