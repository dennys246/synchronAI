"""Build labels.csv for the EmoGrow, R56 P-CAT and R01 P-CAT DB-DOS synchrony codings.

These studies use the same per-second `Time | Code | Notes` coding sheets as
CARE (code `a`=async=0, `s`=sync=1), but two things prevent reuse of
`preprocessing/raw_to_csv.preprocess_raw_to_csv` directly:

1. **Time format.** Coders entered MM:SS ("0:57", "22:17"), which Excel/openpyxl
   store as `datetime.time(minute, second)`. The stock `_parse_time_to_seconds`
   stringifies that to "00:57:00" and reads it as HH:MM:SS -> 57*60=3420s, i.e.
   ~60x too large, pushing every label out of the video's range. We parse MM:SS
   here as `t.hour*60 + t.minute` (openpyxl packs minutes into `.hour`).

2. **Layout.** Coding is rater-major, not `{subject}/{session}/`:
     EmoGrow:  {rater}_Coding/V1/{ID}_Final.xlsx        (one file per participant)
     R56 PCAT: {rater}/{family}_{activity}_Complete.xlsx (activities = sequential,
               non-overlapping absolute-time segments of one {family}_DB-DOS video)
     R01 PCAT: second-by-second/T1/{record}/{record}_T1_{Task}.xlsx (5-digit IDs;
               Time is wall-clock from video start, shared with the verbal VAD export)

Multiple raters are combined by per-second majority vote; ties are dropped.

Video paths are resolved against the LOCAL mac mount for existence/ffprobe, but
emitted with the CLUSTER prefix so the cluster feature-extraction jobs resolve
them. See CLUSTER_ROOT / LOCAL_ROOT below.

Output schema matches labels.csv exactly: video_path,second,label,subject_id,session
"""

from __future__ import annotations

import argparse
import collections
import datetime
import glob
import os
import re
import subprocess

import openpyxl

LOCAL_ROOT = "/Volumes/perlmansusan/Active/moochie"
CLUSTER_ROOT = "/storage1/fs1/perlmansusan/Active/moochie"

# The same NFS share is mounted at different paths on the mac and the cluster.
# Discover which one actually exists rather than assuming the mac: running on
# the cluster used to glob a nonexistent /Volumes path, find zero coding files,
# and write a header-only CSV over a good one without erroring.
DATA_ROOT = next(
    (r for r in (LOCAL_ROOT, CLUSTER_ROOT) if os.path.isdir(f"{r}/study_data")),
    LOCAL_ROOT,
)
STUDY_DATA = f"{DATA_ROOT}/study_data"
LABEL_ENCODING = {"a": 0, "s": 1}


def parse_mmss(val) -> int | None:
    """Coding Time cell -> integer second. openpyxl packs MM:SS into time(MM, SS)."""
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val.hour * 60 + val.minute
    s = str(val).strip()
    if ":" in s:
        p = s.split(":")
        try:
            return int(p[0]) * 60 + int(float(p[1]))
        except (ValueError, TypeError):
            return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def read_coding(path: str) -> list[tuple[int, int]]:
    """Return list of (second, label) from a single Time/Code/Notes sheet."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        print(f"    ! failed to open {os.path.basename(path)}: {e}")
        return []
    ws = wb[wb.sheetnames[0]]
    out: list[tuple[int, int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or row[1] is None:
            continue
        code = str(row[1]).strip().lower()
        if code not in LABEL_ENCODING:
            continue
        sec = parse_mmss(row[0])
        if sec is None:
            continue
        out.append((sec, LABEL_ENCODING[code]))
    wb.close()
    return out


def consensus(rows: list[tuple[int, int]]) -> tuple[dict[int, int], int]:
    """Majority-vote label per second across all raters. Ties dropped."""
    by_sec: dict[int, list[int]] = collections.defaultdict(list)
    for sec, lab in rows:
        by_sec[sec].append(lab)
    labels: dict[int, int] = {}
    ties = 0
    for sec, labs in by_sec.items():
        c = collections.Counter(labs)
        top = c.most_common()
        if len(top) > 1 and top[0][1] == top[1][1]:
            ties += 1
            continue
        labels[sec] = top[0][0]
    return labels, ties


def ffprobe_duration(path: str) -> float | None:
    try:
        out = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=nw=1:nk=1", path],
            capture_output=True, text=True, timeout=120,
        )
        return float(out.stdout.strip())
    except Exception:  # noqa: BLE001
        return None


def emit_path(local_path: str) -> str:
    """Always emit the cluster prefix — extraction jobs run inside Docker there.
    No-op when we are already reading from the cluster mount."""
    return local_path.replace(DATA_ROOT, CLUSTER_ROOT, 1)


# --------------------------------------------------------------------------- #
# EmoGrow
# --------------------------------------------------------------------------- #
def discover_emogrow() -> dict[str, list[str]]:
    """participant ID -> list of coding files across raters."""
    root = f"{STUDY_DATA}/EmoGrow/synchrony_coding/UGRA_Codes/Past RA Coding"
    by_id: dict[str, list[str]] = collections.defaultdict(list)
    for f in glob.glob(os.path.join(root, "*", "**", "*.xlsx"), recursive=True):
        b = os.path.basename(f)
        if b.startswith("~$") or b.startswith("._"):
            continue
        m = re.match(r"(\d{3})[ _]", b)
        if not m:
            continue
        by_id[m.group(1)].append(f)
    return by_id


def resolve_emogrow_video(pid: str) -> str | None:
    """Pick the primary DBDOS cam3 take.

    Filenames are inconsistent ({id}_DBDOS_0_cam3, DBDOS_{id}_0_cam3,
    DBDOS{id}_0_cam3) and some folders also hold retakes (DBDOS2_0_cam3).
    A trailing digit after "DBDOS" that is NOT the participant id (e.g. the "2"
    in DBDOS2) marks a retake and must be de-prioritised -- picking it silently
    mis-aligns every label (caught in QC on 043).
    """
    d = f"{STUDY_DATA}/EmoGrow/video_data/{pid}/V1"
    if not os.path.isdir(d):
        return None
    cands = [
        f for f in os.listdir(d)
        if re.search(r"dbdos.*cam3.*\.(mpg|mp4)$", f, re.I)
        and not f.startswith("._")
    ]
    if not cands:
        return None

    def rank(f: str) -> tuple:
        low = f.lower()
        # retake marker: "dbdos" followed by digits that are not the participant id
        m = re.search(r"dbdos(\d+)", low)
        is_retake = bool(m) and m.group(1).lstrip("0") != pid.lstrip("0")
        primary = "_0_" in low                    # the "_0" take index
        return (is_retake, not primary, f)        # False<True, so non-retake/primary first

    return os.path.join(d, sorted(cands, key=rank)[0])


# --------------------------------------------------------------------------- #
# R56 P-CAT
# --------------------------------------------------------------------------- #
def discover_r56(root_override: str | None = None) -> dict[str, list[str]]:
    """family ID -> list of activity coding files across raters."""
    root = root_override or f"{STUDY_DATA}/P-CAT/R56/synchrony_coding"
    by_id: dict[str, list[str]] = collections.defaultdict(list)
    for f in glob.glob(os.path.join(root, "*", "*.xlsx"), recursive=True):
        b = os.path.basename(f)
        if b.startswith("~$") or b.startswith("._") or b.lower() == "template.xlsx":
            continue
        m = re.match(r"(\d{4})_", b)
        if not m:  # e.g. "BC PCAT Coding.xlsx" summary sheet
            continue
        by_id[m.group(1)].append(f)
    return by_id


def resolve_r56_video(fam: str) -> str | None:
    d = f"{STUDY_DATA}/P-CAT/R56/video_data/{fam}"
    if not os.path.isdir(d):
        return None
    for f in sorted(os.listdir(d)):
        if re.match(rf"{fam}_DB-?DOS\.(mp4|MP4)$", f) and not f.startswith("._"):
            return os.path.join(d, f)
    return None


# --------------------------------------------------------------------------- #
# R01 P-CAT
# --------------------------------------------------------------------------- #
# Time cells here are Excel numFmtId 20 (h:mm) holding what a coder typed as
# m:ss, so openpyxl hands back time(minutes, seconds) and parse_mmss already
# decodes them correctly — verified against an independent day-fraction*1440
# decode on all 21,836 rows of all 22 files, zero mismatches. No new parser.
#
# Only 16 of the 156 subject dirs currently hold coding; the rest are empty
# placeholders awaiting the coding queue. Re-run this as coding lands.
def discover_r01(root_override: str | None = None) -> dict[str, list[str]]:
    """record ID -> list of per-task coding files (Arts / Puzzles / Magnetiles).

    Five-digit IDs, unlike R56's four. Reusing R56's r"(\\d{4})_" here would
    match the first four digits, fail on the trailing digit, and skip every
    file in silence.
    """
    root = root_override or f"{STUDY_DATA}/P-CAT/R01/synchrony_coding/second-by-second/T1"
    by_id: dict[str, list[str]] = collections.defaultdict(list)
    for f in glob.glob(os.path.join(root, "*", "*.xlsx")):
        b = os.path.basename(f)
        if b.startswith("~$") or b.startswith("._"):
            continue
        m = re.match(r"(\d{5})_", b)
        if not m:
            continue
        by_id[m.group(1)].append(f)
    return by_id


def resolve_r01_video(rid: str) -> str | None:
    """Pick the DB-DOS recording, tolerating the known anomalies.

    Rejects AppleDouble sidecars, partial-transfer artifacts (a hash suffix
    after the extension), zero-byte remux failures, and the loudness-fixed
    derivatives, whose audio is heterogeneous down to 8 kHz mono. Accepts the
    'ddbos' typo in 11181. Prefers the largest file when a session was split.
    """
    d = f"{STUDY_DATA}/P-CAT/R01/data/WUSTL_data/T1/video_data/dbdos/{rid}"
    if not os.path.isdir(d):
        return None
    cands = []
    for f in os.listdir(d):
        if f.startswith("._") or "_fixed" in f:
            continue
        if not re.search(r"\.(mp4|mkv)$", f, re.IGNORECASE):
            continue
        if not re.search(r"d[db]dos", f, re.IGNORECASE):
            continue
        p = os.path.join(d, f)
        try:
            size = os.path.getsize(p)
        except OSError:
            continue
        if size:
            cands.append((size, p))
    return max(cands)[1] if cands else None


# --------------------------------------------------------------------------- #
def build(study: str, out_csv: str, no_ffprobe: bool, min_seconds: int) -> None:
    if study == "emogrow":
        by_id = discover_emogrow()
        resolve = resolve_emogrow_video
        sid = lambda i: f"EG{i}"      # noqa: E731
        session = "V1"
    elif study == "r01pcat":
        by_id = discover_r01()
        resolve = resolve_r01_video
        # NEVER a bare 5-digit subject_id: family_id_from_subject
        # (train_multimodal_from_features.py:2120) truncates those to 4 digits,
        # so 11002/11003/11009 would collapse into one CV group. The "-dyad"
        # suffix routes to the earlier branch and round-trips intact.
        sid = lambda i: f"{i}-dyad"   # noqa: E731
        session = "T1"
    else:
        by_id = discover_r56()
        resolve = resolve_r56_video
        sid = lambda i: f"{i}-dyad"   # noqa: E731
        session = "DBDOS"

    rows_out: list[tuple[str, int, int, str, str]] = []
    print(f"\n=== {study}: {len(by_id)} coded IDs (min_seconds={min_seconds}) ===")
    print(f"    reading coding from {STUDY_DATA}")
    if not by_id:
        # Refuse to write. Silently emitting a header-only CSV here overwrote a
        # good labels file once, because a wrong data root globs to nothing and
        # looks identical to "this study has no coding yet".
        raise SystemExit(
            f"ERROR: found 0 coding files for '{study}' under {STUDY_DATA}.\n"
            f"       Check that path exists and holds the expected layout. "
            f"Refusing to write {out_csv}."
        )
    n_ok = n_novideo = n_sparse = 0
    dist = collections.Counter()
    for pid in sorted(by_id):
        files = by_id[pid]
        rows = []
        for f in files:
            rows.extend(read_coding(f))
        labels, ties = consensus(rows)
        vid_local = resolve(pid)
        if vid_local is None:
            print(f"  {pid}: {len(labels):4d} sec, {len(files)} raters -> NO VIDEO (skipped)")
            n_novideo += 1
            continue
        dur = None if no_ffprobe else ffprobe_duration(vid_local)
        vid_cluster = emit_path(vid_local)
        buf: list[tuple[str, int, int, str, str]] = []
        oob = 0
        for sec in sorted(labels):
            if dur is not None and sec + 1 > dur:
                oob += 1
                continue
            buf.append((vid_cluster, sec, labels[sec], sid(pid), session))
        kept = len(buf)
        durs = f"{dur:.0f}s" if dur else "n/a"
        flag = f" OOB={oob}" if oob else ""
        tie = f" ties={ties}" if ties else ""
        if kept < min_seconds:
            print(f"  {pid}: {kept:4d} sec -> SPARSE (< {min_seconds}, dropped){tie} "
                  f"{os.path.basename(vid_local)}")
            n_sparse += 1
            continue
        rows_out.extend(buf)
        for _, _, lab, _, _ in buf:
            dist[lab] += 1
        print(f"  {pid}: {kept:4d} sec kept (raters={len(files)}, vid_dur={durs}{flag}{tie}) "
              f"{os.path.basename(vid_local)}")
        n_ok += 1

    os.makedirs(os.path.dirname(os.path.abspath(out_csv)), exist_ok=True)
    with open(out_csv, "w") as fh:
        fh.write("video_path,second,label,subject_id,session\n")
        for vp, sec, lab, s, sess in rows_out:
            # video paths here never contain commas
            fh.write(f"{vp},{sec},{lab},{s},{sess}\n")

    total = sum(dist.values())
    print(f"\n  participants kept: {n_ok}   sparse-dropped: {n_sparse}   no-video: {n_novideo}")
    if total:
        print(f"  labeled seconds: {total}  (async={dist[0]}={100*dist[0]/total:.1f}%  "
              f"sync={dist[1]}={100*dist[1]/total:.1f}%)")
    print(f"  wrote {out_csv}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--study", required=True, choices=["emogrow", "r56pcat", "r01pcat"])
    ap.add_argument("--out", required=True)
    ap.add_argument("--no-ffprobe", action="store_true",
                    help="skip video-duration probe (faster, no OOB filtering)")
    ap.add_argument("--min-seconds", type=int, default=120,
                    help="drop participants with fewer than this many kept labeled seconds")
    args = ap.parse_args()
    build(args.study, args.out, args.no_ffprobe, args.min_seconds)


if __name__ == "__main__":
    main()
