"""CARE global (block-level) synchrony code ingest.

Layout: study_data/CARE/synchrony_coding/global participant codes/
            {V0,V1,V2}/{5-digit child id}/{id}_{visit}_{Activity}[_N].xlsx

One file = one coded block. Fixed template (both generations share cell
positions; only sharedStrings order and the AVERAGE SYNCHORNY/SYNCHRONY
spelling differ): header row 4, data rows 5-8 --
    B5 TIMEPOINT | C5 ID (numeric) | D5 BLOCK activity | E5:E8 "TRIAL 0".."TRIAL 3"
    F5:F8 SCORE (0-5 in 0.5 steps) | G5 =AVERAGE(F5:F8)
B/C/D are merged over rows 5-8 so only row 5 carries them. The _N filename
suffix is the reliability coder index, not a block. Filenames are messy
(4-digit ID typos, stray spaces, Arts/Art/Arts&Crafts vs Magnetiles/Magnets/
MagnetTiles), so the participant DIRECTORY is the identity authority and the
in-file D5 cell is the activity authority; both are cross-checked and logged.
"""

from __future__ import annotations

import collections
import re
from pathlib import Path

import openpyxl

ACTIVITY_TO_BLOCK = {"arts": 1, "puzzles": 2, "magnetiles": 3}


def normalize_activity(s: str) -> str | None:
    t = re.sub(r"[^a-z]", "", str(s).lower())
    if t.startswith("art"):
        return "arts"
    if t.startswith("puzzle"):
        return "puzzles"
    if t.startswith(("magnet", "mt")):
        return "magnetiles"
    return None


def parse_global_code_file(path: str | Path) -> dict | None:
    """One block file -> {id_cell, activity, scores: {trial: score}, average}.
    Returns None (caller logs) when the sheet doesn't match the template."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        cells = {f"{c}{r}": ws.cell(row=r, column=col).value
                 for col, c in ((2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G"))
                 for r in range(4, 9)}
    finally:
        wb.close()
    header_f = str(cells.get("F4") or "").strip().upper()
    if header_f != "SCORE":
        return None
    scores: dict[int, float] = {}
    for r in range(5, 9):
        trial_lbl = str(cells.get(f"E{r}") or "")
        m = re.search(r"TRIAL\s*(\d)", trial_lbl, re.IGNORECASE)
        val = cells.get(f"F{r}")
        if m and isinstance(val, (int, float)):
            scores[int(m.group(1))] = float(val)
    if not scores:
        return None
    avg = cells.get("G5")
    return {
        "id_cell": str(cells.get("C5") or "").split(".")[0],
        "timepoint_cell": str(cells.get("B5") or "").strip(),
        "activity": normalize_activity(cells.get("D5") or ""),
        "scores": scores,
        "average": float(avg) if isinstance(avg, (int, float)) else None,
    }


def _coder_index(filename: str) -> int:
    m = re.search(r"[_ ](\d)\.xlsx$", filename)
    return int(m.group(1)) if m else 1


def parse_global_codes(root: str | Path, limit: int = 0) -> tuple[list[dict], list[str]]:
    """All coder-level trial scores under the global-codes root.

    Returns (rows, problems). Row: visit, subject_id, family_id, activity,
    block, coder, trial, score. `problems` collects skipped/mismatched files.
    """
    root = Path(root)
    rows: list[dict] = []
    problems: list[str] = []
    n_files = 0
    for visit_dir in sorted(root.iterdir()):
        if not re.fullmatch(r"V\d", visit_dir.name) or not visit_dir.is_dir():
            continue
        for sub_dir in sorted(visit_dir.iterdir()):
            if not re.fullmatch(r"\d{5}", sub_dir.name) or not sub_dir.is_dir():
                continue
            for f in sorted(sub_dir.glob("*.xlsx")):
                if f.name.startswith(("~$", "._")):
                    continue
                if limit and n_files >= limit:
                    return rows, problems
                n_files += 1
                rel = f"{visit_dir.name}/{sub_dir.name}/{f.name}"
                try:
                    parsed = parse_global_code_file(f)
                except Exception as e:  # noqa: BLE001 - openpyxl raises broadly
                    problems.append(f"{rel}: read failed: {e}")
                    continue
                if parsed is None:
                    problems.append(f"{rel}: does not match block template")
                    continue
                if parsed["activity"] is None:
                    # BLOCK cell unusable (template placeholder, odd label);
                    # fall back to the activity token in the filename.
                    m = re.search(r"_V\d_(.+?)(?:[_ ]\d)?\.xlsx$", f.name)
                    fallback = normalize_activity(m.group(1)) if m else None
                    if fallback is None:
                        problems.append(f"{rel}: unrecognized BLOCK {parsed!r}")
                        continue
                    problems.append(f"{rel}: BLOCK cell unusable, using filename "
                                    f"activity '{fallback}'")
                    parsed["activity"] = fallback
                if parsed["id_cell"] and parsed["id_cell"] != sub_dir.name:
                    problems.append(f"{rel}: ID cell {parsed['id_cell']} != dir "
                                    f"{sub_dir.name} (using dir)")
                for trial, score in sorted(parsed["scores"].items()):
                    rows.append({
                        "visit": visit_dir.name,
                        "subject_id": sub_dir.name,
                        "family_id": sub_dir.name[:4],
                        "activity": parsed["activity"],
                        "block": ACTIVITY_TO_BLOCK[parsed["activity"]],
                        "coder": _coder_index(f.name),
                        "trial": trial,
                        "score": score,
                        "source_file": rel,
                    })
    return rows, problems


def consensus_scores(rows: list[dict]) -> list[dict]:
    """Mean across coders per (visit, subject, activity, trial)."""
    grouped: dict[tuple, list[float]] = collections.defaultdict(list)
    for r in rows:
        grouped[(r["visit"], r["subject_id"], r["activity"], r["block"],
                 r["trial"])].append(r["score"])
    out = []
    for (visit, sub, act, block, trial), scores in sorted(grouped.items()):
        out.append({
            "visit": visit, "subject_id": sub, "family_id": sub[:4],
            "activity": act, "block": block, "trial": trial,
            "score_mean": sum(scores) / len(scores),
            "n_coders": len(scores),
            "score_spread": max(scores) - min(scores),
        })
    return out
